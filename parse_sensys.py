#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
import argparse
import sys
import io
import re

def illegal_char_remover(data):
    ILLEGAL_CHARACTERS_RE = re.compile(
        r'[\000-\010]|[\013-\014]|[\016-\037]|[\x00-\x1f\x7f-\x9f]|[\uffff]')
    """Remove ILLEGAL CHARACTER."""
    if isinstance(data, str):
        return ILLEGAL_CHARACTERS_RE.sub("", data)
    else:
        return data

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

parser = argparse.ArgumentParser(description="excel hoge hoge")
parser.add_argument("filename", help="filename hog hoge")

args = parser.parse_args()
filename = args.filename

fp = open(filename, "r", encoding="utf-8")

text = fp.read()

index = 0

index_start = text.find("<li>", index)

wb = openpyxl.Workbook()
ws = wb.active

# print only value is set
i = 1
ret = ws.cell(row=i, column=1).value
while ret != None:
    print(ret)
    i = i + 1
    ret = ws.cell(row=i, column=1).value


row_excel = 1
    
while index_start != -1:
    index_end = text.find("</li>", index_start)

#    print(index_start)
#    print(index_end)

    print(text[index_start+4:index_end])
    target = text[index_start+4:index_end]
    target = target.strip()
    end_title = target.find("<br>")
    if end_title == -1:
        index_start = text.find("<li>", index_end)
        row_excel = row_excel + 1
        continue

    title = target[0:end_title]
    title = title.strip()
    p = re.compile(r"<[^>]*?>")
    title = p.sub("", title)
    print(title)
    start_author = target.find('<span class="authors">')
    start_author += len('<span class="authors">')
    end_author = target.find('</span>', start_author)
    text_author = target[start_author:end_author]
    text_author = illegal_char_remover(text_author)
    print(text_author)
    ws.cell(row=row_excel, column=1).value = title
    ws.cell(row=row_excel, column=2).value = text_author
    index_start = text.find("<li>", index_end)
    row_excel = row_excel + 1


wb.save(filename + ".xlsx")    
