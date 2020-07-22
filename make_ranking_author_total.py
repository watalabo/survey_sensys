#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
import argparse
import sys
import io
import re

def illegal_char_remover(data):
    ILLEGAL_CHARACTERS_RE = re.compile(
        r'[\000-\010]|[\013-\014]|[\016-\037]|[\x00-\x1f\x7f-\x9f]|[\uffff]')
    """Remove ILLEGAL CHARACTER."""
    if isinstance(data, str):
        return ILLEGAL_CHARACTERS_RE.sub("", data)
    else:
        return data


def make_author_db():
    fp = open("authordb.txt", "r", encoding="cp932")
    lines = fp.read().split("\n")
    db = set()
    for line in lines:
        db.add(line.strip())
    return db


def count_author(author, filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    i = 1
    count = 0
    ret = ws.cell(row=i, column=2).value
    while ret != None:
#        print(ret)
        line = ret
        i = i + 1
        line = line.replace(";", ",")
        line = line.replace(" and ", ",")
        line = line.replace("David E. Culler", "David Culler")
        line = line.replace("  ", " ")
        line = line.replace("  ", " ")
        line = line.replace('"', '')
        line = line.replace('"', '')
        line = line.replace('Kay Roemer', 'Kay Romer')
        line = line.replace('Kaushik R Chowdhury', 'Kaushik Chowdhury')
        line = line.replace('Richard E. Howard', 'Richard Howard')

        if line.find(author) >= 0:
            count = count + 1
        ret = ws.cell(row=i, column=2).value

    return count

authordb = make_author_db()

ranking = {}

for author in authordb:
#    xlsxlst = ["2019_sensys.txt.xlsx","2018_sensys.txt.xlsx","2017_sensys.txt.xlsx","2016_sensys.txt.xlsx","2015_sensys.txt.xlsx"]
    xlsxlst = ["2019_sensys.txt.xlsx","2018_sensys.txt.xlsx"]
    count = 0
    for filename in xlsxlst:
        ret = count_author(author, filename)
        sys.stdout.flush()
        count = count + ret
#    print("%s, %d" % (author, count))
    ranking[author] = count
        

ranking = sorted(ranking.items(), key=lambda x:x[1], reverse=True)

for items in ranking:
#    print(items)
    print("%s, %d" % (items[0], items[1]))
    
